import requests
import json
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Color
from bs4 import BeautifulSoup


"""--------------------------------------Extract Link List--------------------------------------"""


def get_book_links(book):
    atag = book.find("a")
    href = atag["href"]
    link = "https://www.barnesandnoble.com" + href.split(";")[0]
    return link


def get_page_info(page):
    result = requests.get("https://www.barnesandnoble.com/b/books/_/N-1fZ29Z8q8?Nrpp=20&page=" + str(page))
    bs_obj = BeautifulSoup(result.content, "html.parser")
    shelf = bs_obj.find("div", {"class": "product-shelf-list"})
    books = shelf.findAll("h3", {"class": "product-info-title"})
    link_list = [get_book_links(book) for book in books]
    return link_list


result = []

for page in range(1, 6):
    result.extend(get_page_info(page))


def get_links():
    return result


"""---------------------------------Extract Book Information---------------------------------"""


def get_book_info(url):
    result = requests.get(url)
    bs_obj = BeautifulSoup(result.content, "html.parser")

    table = bs_obj.find("table", {"class": "plain centered"})
    table_detail = table.findAll("td")
    contributors = bs_obj.find("span", {"class": "contributors"})
    publisher_blank = table_detail[1].text.split("\n")

    title = bs_obj.find("h1", {"class": "pdp-header-title"}).text
    author = contributors.find("a").text
    publisher = publisher_blank[1]
    cover = bs_obj.find("h2", {"id": "pdp-info-format"}).text
    isbn = table_detail[0].text
    date = table_detail[2].text

    # for books not on sale
    try:
        old_price = bs_obj.find("s", {"class": "old-price"}).text
        price = old_price
    except:
        current_price = bs_obj.find("span", {"class": "price current-price ml-0"}).text
        price = current_price

    dictionary1 = {}
    dictionary1["title"] = title.replace("'", '"')
    dictionary1["author"] = author.replace("'", '"')
    dictionary1["publisher"] = publisher.replace("'", '"')
    dictionary1["price"] = price
    dictionary1["cover"] = cover.split("\n")[0]
    dictionary1["ISBN"] = isbn
    dictionary1["date"] = date

    return dictionary1


urls = get_links()


def get_info():
    book_info = [get_book_info(url) for url in urls]
    return book_info


"""--------------------------------------Save in Excel--------------------------------------"""


j1 = get_info()
j2 = json.dumps(j1)
j3 = json.loads(j2)

wb = Workbook()
ws1 = wb.active
ws1.freeze_panes = "A2"

ws1["A1"] = "제목"
ws1["A1"].font = Font(size=12, bold=True)
ws1["A1"].fill = PatternFill(patternType="solid", fgColor=Color("B8CCE4"))
ws1["B1"] = "저자"
ws1["B1"].font = Font(size=12, bold=True)
ws1["B1"].fill = PatternFill(patternType="solid", fgColor=Color("B8CCE4"))
ws1["C1"] = "출판사"
ws1["C1"].font = Font(size=12, bold=True)
ws1["C1"].fill = PatternFill(patternType="solid", fgColor=Color("B8CCE4"))
ws1["D1"] = "가격"
ws1["D1"].font = Font(size=12, bold=True)
ws1["D1"].fill = PatternFill(patternType="solid", fgColor=Color("B8CCE4"))
ws1["E1"] = "표지"
ws1["E1"].font = Font(size=12, bold=True)
ws1["E1"].fill = PatternFill(patternType="solid", fgColor=Color("B8CCE4"))
ws1["F1"] = "ISBN"
ws1["F1"].font = Font(size=12, bold=True)
ws1["F1"].fill = PatternFill(patternType="solid", fgColor=Color("B8CCE4"))
ws1["G1"] = "출판일"
ws1["G1"].font = Font(size=12, bold=True)
ws1["G1"].fill = PatternFill(patternType="solid", fgColor=Color("B8CCE4"))


for row_index in range(1, len(j1)+1):
    ws1.cell(row=row_index + 1, column=1).value = j3[row_index - 1]["title"]
    ws1.cell(row=row_index + 1, column=2).value = j3[row_index - 1]["author"]
    ws1.cell(row=row_index + 1, column=3).value = j3[row_index - 1]["publisher"]
    ws1.cell(row=row_index + 1, column=4).value = j3[row_index - 1]["price"]
    ws1.cell(row=row_index + 1, column=5).value = j3[row_index - 1]["cover"]
    ws1.cell(row=row_index + 1, column=6).value = j3[row_index - 1]["ISBN"]
    ws1.cell(row=row_index + 1, column=7).value = j3[row_index - 1]["date"]

wb.save("C:\\Users\\user\\Desktop\\우제현\\Barnes_And_Noble_Top100.xlsx")
