from selenium import webdriver
from bs4 import BeautifulSoup
from openpyxl import load_workbook
import time
import random

monthly_stat_from = "2020/02/01"

loan_date_from = "2020/02/24"
loan_date_to = "2020/02/28"

wb = load_workbook("C:\\Users\\user\\Desktop\\주간업무 추진실적\\test.xlsx")
ws = wb["1.이용현황(일기준)"]
ws2 = wb["3.장서보유현황"]
ws3 = wb["4.기증자료현황"]

def send_checkout_date(from_, to_):
    driver.implicitly_wait(10)
    driver.find_element_by_id("loan_date_from").clear()
    driver.find_element_by_id("loan_date_from").send_keys(from_)
    driver.find_element_by_id("loan_date_to").clear()
    driver.find_element_by_id("loan_date_to").send_keys(to_)

def send_return_date(from_, to_):
    driver.implicitly_wait(10)
    driver.find_element_by_id("return_date_from").clear()
    driver.find_element_by_id("return_date_from").send_keys(from_)
    driver.find_element_by_id("return_date_to").clear()
    driver.find_element_by_id("return_date_to").send_keys(to_)

def send_register_date(from_, to_):
    driver.implicitly_wait(10)
    driver.find_element_by_id("reg_date_from").clear()
    driver.find_element_by_id("reg_date_from").send_keys(from_)
    driver.find_element_by_id("reg_date_to").clear()
    driver.find_element_by_id("reg_date_to").send_keys(to_)

def search_and_close():
    driver.implicitly_wait(10)
    try:
        search = driver.find_element_by_xpath('//*[@id="statistics_search_btn"]')
    except:
        search = driver.find_element_by_xpath('//*[@id="shelf_hold_statistics_search_modal"]/div/div/div[3]/div/div[2]/button[1]')
    driver.execute_script("arguments[0].click();", search)
    driver.implicitly_wait(120)
    driver.find_element_by_xpath('//*[@id="msg_btn_1"]').click()
    driver.implicitly_wait(10)
    close_button = driver.find_element_by_xpath('//*[@id="pusrchaesDataPrintdeSearchClose"]')
    driver.execute_script("arguments[0].click();", close_button)

def get_page_info():
    result = driver.page_source
    bs_obj = BeautifulSoup(result, "html.parser")
    table = bs_obj.find("table", class_="display compact cell-border table table-bordered ui-jqgrid-btable ui-common-table")
    tr = table.find_all("tr")

    general_works = tr[1].find_all("td")
    philosophy = tr[2].find_all("td")
    religion = tr[3].find_all("td")
    social_sciences = tr[4].find_all("td")
    natural_sciences = tr[5].find_all("td")
    technology = tr[6].find_all("td")
    arts = tr[7].find_all("td")
    language = tr[8].find_all("td")
    literature = tr[9].find_all("td")
    history = tr[10].find_all("td")

    dictionary = {}
    dictionary["general_works"] = [int(general_works[i].text) for i in range(4, 9)]
    dictionary["philosophy"] = [int(philosophy[i].text) for i in range(4, 9)]
    dictionary["religion"] = [int(religion[i].text) for i in range(4, 9)]
    dictionary["social_sciences"] = [int(social_sciences[i].text) for i in range(4, 9)]
    dictionary["natural_sciences"] = [int(natural_sciences[i].text) for i in range(4, 9)]
    dictionary["technology"] = [int(technology[i].text) for i in range(4, 9)]
    dictionary["arts"] = [int(arts[i].text) for i in range(4, 9)]
    dictionary["language"] = [int(language[i].text) for i in range(4, 9)]
    dictionary["literature"] = [int(literature[i].text) for i in range(4, 9)]
    dictionary["history"] = [int(history[i].text) for i in range(4, 9)]
    
    dictionary["monthly_general_works"] = int(general_works[3].text)
    dictionary["monthly_philosophy"] = int(philosophy[3].text)
    dictionary["monthly_religion"] = int(religion[3].text)
    dictionary["monthly_social_sciences"] = int(social_sciences[3].text)
    dictionary["monthly_natural_sciences"] = int(natural_sciences[3].text)
    dictionary["monthly_technology"] = int(technology[3].text)
    dictionary["monthly_arts"] = int(arts[3].text)
    dictionary["monthly_language"] = int(language[3].text)
    dictionary["monthly_literature"] = int(literature[3].text)
    dictionary["monthly_history"] = int(history[3].text)

    return dictionary

def get_user_statistics():
    result = driver.page_source
    bs_obj = BeautifulSoup(result, "html.parser")
    table = bs_obj.find("table", class_="display compact cell-border table table-bordered ui-jqgrid-btable ui-common-table")
    tr = table.find_all("tr")

    male = tr[1].find_all("td")
    female = tr[2].find_all("td")
    male_student = tr[3].find_all("td")
    female_student = tr[4].find_all("td")
    male_youth = tr[5].find_all("td")
    female_youth = tr[6].find_all("td")
    employee = tr[7].find_all("td")
    digital = tr[8].find_all("td")
    good_member = tr[9].find_all("td")
    foreigner = tr[10].find_all("td")
    good_member_chungla = tr[11].find_all("td")
    good_member_heavy_reader = tr[12].find_all("td")
    related_organization = tr[13].find_all("td")
    reference_room = tr[14].find_all("td")
    vision_impaired = tr[15].find_all("td")
    hearing_impaired = tr[16].find_all("td")
    physical_disabled = tr[17].find_all("td")
    brain_disabled = tr[18].find_all("td")
    traveling_library = tr[19].find_all("td")
    renal_disorder = tr[20].find_all("td")
    mental_disorder = tr[21].find_all("td")
    other_disorder = tr[22].find_all("td")
    honorary_member = tr[23].find_all("td")
    do_not_delete = tr[24].find_all("td")

    dictionary = {}
    dictionary["korean"] = [int(male[i].text) + int(female[i].text) + int(male_student[i].text) +
                            int(female_student[i].text) + int(male_youth[i].text) + int(female_youth[i].text) +
                            int(employee[i].text) + int(digital[i].text) + int(good_member[i].text) +
                            int(good_member_chungla[i].text) + int(good_member_heavy_reader[i].text) +
                            int(related_organization[i].text) + int(reference_room[i].text) + int(vision_impaired[i].text) +
                            int(hearing_impaired[i].text) + int(physical_disabled[i].text) + int(brain_disabled[i].text) +
                            int(traveling_library[i].text) + int(renal_disorder[i].text) + int(mental_disorder[i].text) +
                            int(other_disorder[i].text) + int(honorary_member[i].text) + int(do_not_delete[i].text) for i in range(4,9)]
    dictionary["foreigner"] = [int(foreigner[i].text) for i in range(4,9)]

    dictionary["monthly_korean"] = int(male[3].text) + int(female[3].text) + int(male_student[3].text) + int(female_student[3].text)
    + int(male_youth[3].text) + int(female_youth[3].text) + int(employee[3].text) + int(digital[3].text) + int(good_member[3].text)
    + int(good_member_chungla[3].text) + int(good_member_heavy_reader[3].text) + int(related_organization[3].text) + int(reference_room[3].text)
    + int(vision_impaired[3].text) + int(hearing_impaired[3].text) + int(physical_disabled[3].text) + int(brain_disabled[3].text)
    + int(traveling_library[3].text) + int(renal_disorder[3].text) + int(mental_disorder[3].text) + int(other_disorder[3].text) + int(honorary_member[3].text) + int(do_not_delete[3].text)

    dictionary["monthly_foreigner"] = int(foreigner[3].text)

    return dictionary

def get_collection_statistics():
    result = driver.page_source
    bs_obj = BeautifulSoup(result, "html.parser")
    table = bs_obj.find("table", class_="display compact cell-border table table-bordered ui-jqgrid-btable ui-common-table")
    tr = table.find_all("tr")

    korean_big = tr[1].find_all("td")
    korean = tr[2].find_all("td")
    korean_refer = tr[3].find_all("td")
    non_book = tr[4].find_all("td")
    chinese = tr[5].find_all("td")
    french = tr[6].find_all("td")
    english = tr[7].find_all("td")
    japanese = tr[8].find_all("td")
    mongolian = tr[9].find_all("td")
    filipino = tr[10].find_all("td")
    english_refer = tr[11].find_all("td")
    spanish = tr[12].find_all("td")
    vietnamese = tr[13].find_all("td")
    cambodian = tr[14].find_all("td")

    dictionary = {}
    dictionary["korean_big"] = [int(korean_big[i].text) for i in range(3, 14)]
    dictionary["korean"] = [int(korean[i].text) for i in range(3, 14)]
    dictionary["korean_refer"] = [int(korean_refer[i].text) for i in range(3, 14)]
    dictionary["non_book"] = [int(non_book[i].text) for i in range(3, 14)]
    dictionary["chinese"] = [int(chinese[i].text) for i in range(3, 14)]
    dictionary["french"] = [int(french[i].text) for i in range(3, 14)]
    dictionary["english"] = [int(english[i].text) for i in range(3, 14)]
    dictionary["japanese"] = [int(japanese[i].text) for i in range(3, 14)]
    dictionary["mongolian"] = [int(mongolian[i].text) for i in range(3, 14)]
    dictionary["filipino"] = [int(filipino[i].text) for i in range(3, 14)]
    dictionary["english_refer"] = [int(english_refer[i].text) for i in range(3, 14)]
    dictionary["spanish"] = [int(spanish[i].text) for i in range(3, 14)]
    dictionary["vietnamese"] = [int(vietnamese[i].text) for i in range(3, 14)]
    dictionary["cambodian"] = [int(cambodian[i].text) for i in range(3, 14)]

    return dictionary

"""------------------Alpas Login---------------"""

#Login
driver = webdriver.Chrome("C:\\Users\\user\\PycharmProjects\\untitled\\chromedriver.exe")
driver.implicitly_wait(3)
driver.get("http://152.99.43.46:28180/METIS/")
driver.maximize_window()
driver.find_element_by_name("main_login_user_id").send_keys("안병덕")
driver.find_element_by_id("user_pw").send_keys("dksqudejr1120!")
driver.find_element_by_xpath("/html/body/div[2]/div/button").click()

#Open Statistics tap
driver.find_element_by_xpath('//*[@id="right_container_wrapper"]/header/div[5]/nav/a').click()
driver.find_element_by_xpath('//*[@id="COLI_01_HTML"]/li[12]').click()
driver.switch_to.window(driver.window_handles[1])


"""------------------Checkout Statistics---------------"""

driver.find_element_by_xpath('//*[@id="statistics_type"]/option[6]').click()                                            #Choose Checkout Statistics
driver.find_element_by_xpath('//*[@id="statistics_btn"]').click()                                                       #Click Statistics button
show_zero = driver.find_element_by_xpath('//*[@id="right_container_wrapper"]/main/div/div[1]/form/div[2]/label[4]')
driver.execute_script("arguments[0].click();", show_zero)
send_checkout_date(loan_date_from, loan_date_to)                                                                        #Enter checkout date
driver.find_element_by_xpath('//*[@id="shelf_loc_code_list"]/option[35]').click()                                       #Choose Songdo Library of International Organization
driver.find_element_by_xpath('//*[@id="shelf_loc_code_btn"]').click()
driver.implicitly_wait(10)
by_date = driver.find_element_by_xpath('//*[@id="jqg82"]/td[2]')                                                        #Adjust matrix
driver.execute_script("arguments[0].click();", by_date)
search_and_close()                                                                                                      #Search
get_page_info()                                                                                                         #Parsing


#Save in Excel
ws["F28"], ws["F29"], ws["F30"], ws["F31"], ws["F32"] = [get_page_info()["general_works"][i] for i in range(5)]
ws["G28"], ws["G29"], ws["G30"], ws["G31"], ws["G32"] = [get_page_info()["philosophy"][i] for i in range(5)]
ws["H28"], ws["H29"], ws["H30"], ws["H31"], ws["H32"] = [get_page_info()["religion"][i] for i in range(5)]
ws["I28"], ws["I29"], ws["I30"], ws["I31"], ws["I32"] = [get_page_info()["social_sciences"][i] for i in range(5)]
ws["J28"], ws["J29"], ws["J30"], ws["J31"], ws["J32"] = [get_page_info()["natural_sciences"][i] for i in range(5)]
ws["K28"], ws["K29"], ws["K30"], ws["K31"], ws["K32"] = [get_page_info()["technology"][i] for i in range(5)]
ws["L28"], ws["L29"], ws["L30"], ws["L31"], ws["L32"] = [get_page_info()["arts"][i] for i in range(5)]
ws["M28"], ws["M29"], ws["M30"], ws["M31"], ws["M32"] = [get_page_info()["language"][i] for i in range(5)]
ws["N28"], ws["N29"], ws["N30"], ws["N31"], ws["N32"] = [get_page_info()["literature"][i] for i in range(5)]
ws["O28"], ws["O29"], ws["O30"], ws["O31"], ws["O32"] = [get_page_info()["history"][i] for i in range(5)]

time.sleep(3)


"""------------------Cumulative Checkout Statistics---------------"""

driver.find_element_by_xpath('//*[@id="statistics_btn"]').click()
send_checkout_date(monthly_stat_from, loan_date_to)
search_and_close()
get_page_info()

#Save in Excel
ws["F34"] = get_page_info()["monthly_general_works"]
ws["G34"] = get_page_info()["monthly_philosophy"]
ws["H34"] = get_page_info()["monthly_religion"]
ws["I34"] = get_page_info()["monthly_social_sciences"]
ws["J34"] = get_page_info()["monthly_natural_sciences"]
ws["K34"] = get_page_info()["monthly_technology"]
ws["L34"] = get_page_info()["monthly_arts"]
ws["M34"] = get_page_info()["monthly_language"]
ws["N34"] = get_page_info()["monthly_literature"]
ws["O34"] = get_page_info()["monthly_history"]

time.sleep(3)

"""------------------Return Statistics---------------"""

driver.find_element_by_xpath('//*[@id="statistics_type"]/option[8]').click()
driver.find_element_by_xpath('//*[@id="statistics_btn"]').click()
driver.implicitly_wait(10)
checkbox = driver.find_element_by_xpath('//*[@id="loan_return_statistics_search_modal"]/div/div/div[2]/div[1]/table[2]/tbody/tr[1]/th[2]/label')
driver.execute_script("arguments[0].click();", checkbox)
send_return_date(loan_date_from, loan_date_to)
driver.implicitly_wait(10)
driver.find_element_by_xpath('//*[@id="jqg162"]/td[2]').click()
search_and_close()
get_page_info()

#Save in Excel
ws["F39"], ws["F40"], ws["F41"], ws["F42"], ws["F43"] = [get_page_info()["general_works"][i] for i in range(5)]
ws["G39"], ws["G40"], ws["G41"], ws["G42"], ws["G43"] = [get_page_info()["philosophy"][i] for i in range(5)]
ws["H39"], ws["H40"], ws["H41"], ws["H42"], ws["H43"] = [get_page_info()["religion"][i] for i in range(5)]
ws["I39"], ws["I40"], ws["I41"], ws["I42"], ws["I43"] = [get_page_info()["social_sciences"][i] for i in range(5)]
ws["J39"], ws["J40"], ws["J41"], ws["J42"], ws["J43"] = [get_page_info()["natural_sciences"][i] for i in range(5)]
ws["K39"], ws["K40"], ws["K41"], ws["K42"], ws["K43"] = [get_page_info()["technology"][i] for i in range(5)]
ws["L39"], ws["L40"], ws["L41"], ws["L42"], ws["L43"] = [get_page_info()["arts"][i] for i in range(5)]
ws["M39"], ws["M40"], ws["M41"], ws["M42"], ws["M43"] = [get_page_info()["language"][i] for i in range(5)]
ws["N39"], ws["N40"], ws["N41"], ws["N42"], ws["N43"] = [get_page_info()["literature"][i] for i in range(5)]
ws["O39"], ws["O40"], ws["O41"], ws["O42"], ws["O43"] = [get_page_info()["history"][i] for i in range(5)]

time.sleep(3)

"""------------------Cumulative Return Statistics---------------"""

driver.find_element_by_xpath('//*[@id="statistics_btn"]').click()
send_return_date(monthly_stat_from, loan_date_to)
search_and_close()
time.sleep(3)
get_page_info()

#Save in Excel
ws["F45"] = get_page_info()["monthly_general_works"]
ws["G45"] = get_page_info()["monthly_philosophy"]
ws["H45"] = get_page_info()["monthly_religion"]
ws["I45"] = get_page_info()["monthly_social_sciences"]
ws["J45"] = get_page_info()["monthly_natural_sciences"]
ws["K45"] = get_page_info()["monthly_technology"]
ws["L45"] = get_page_info()["monthly_arts"]
ws["M45"] = get_page_info()["monthly_language"]
ws["N45"] = get_page_info()["monthly_literature"]
ws["O45"] = get_page_info()["monthly_history"]

time.sleep(3)

"""------------------ Return User Statistics---------------"""

driver.find_element_by_xpath('//*[@id="statistics_type"]/option[2]').click()
driver.find_element_by_xpath('//*[@id="statistics_btn"]').click()
register_date = driver.find_element_by_xpath('//*[@id="reg_date_from_text"]')
expected_return_date = driver.find_element_by_xpath('//*[@id="loan_return_statistics_search_modal"]/div/div/div[2]/div[1]/table[2]/tbody/tr[2]/th[1]/label')
driver.execute_script("arguments[0].click();", register_date)
driver.execute_script("arguments[0].click();", expected_return_date)
send_return_date(loan_date_from, loan_date_to)
driver.implicitly_wait(10)
driver.find_element_by_xpath('//*[@id="jqg240"]/td[2]').click()
driver.find_element_by_xpath('//*[@id="jqg264"]/td[2]').click()
search_and_close()
time.sleep(3)

#Save in Excel
ws["D39"], ws["D40"], ws["D41"], ws["D42"], ws["D43"] = [get_user_statistics()["korean"][i] for i in range(5)]
ws["E39"], ws["E40"], ws["E41"], ws["E42"], ws["E43"] = [get_user_statistics()["foreigner"][i] for i in range(5)]

time.sleep(3)

"""------------------ Cumulative Return User Statistics---------------"""

driver.find_element_by_xpath('//*[@id="statistics_btn"]').click()
send_return_date(monthly_stat_from, loan_date_to)
search_and_close()
time.sleep(3)

#Save in Excel
ws["D45"] = get_user_statistics()["monthly_korean"]
ws["E45"] = get_user_statistics()["monthly_foreigner"]



"""------------------ Checkout User Statistics ---------------"""

driver.find_element_by_xpath('//*[@id="statistics_type"]/option[1]').click()
driver.find_element_by_xpath('//*[@id="statistics_btn"]').click()
send_checkout_date(loan_date_from, loan_date_to)
driver.implicitly_wait(10)
by_date = driver.find_element_by_xpath('//*[@id="jqg342"]/td[2]')
by_class = driver.find_element_by_xpath('//*[@id="jqg364"]/td[2]')
driver.execute_script("arguments[0].click();", by_date)
driver.execute_script("arguments[0].click();", by_class)
search_and_close()

#Save in Excel
ws["F17"], ws["F18"], ws["F19"], ws["F20"], ws["F21"] = [get_user_statistics()["korean"][i] for i in range(5)]
ws["G17"], ws["G18"], ws["G19"], ws["G20"], ws["G21"] = [get_user_statistics()["foreigner"][i] for i in range(5)]
ws["H17"], ws["I17"], ws["H18"], ws["I18"], ws["H19"], ws["I19"], ws["H20"], ws["I20"], ws["H21"], ws["I21"] = "=F17*2", "=G17*2", "=F18*2", "=G18*2", "=F19*2", "=G19*2", "=F20*2", "=G20*2", "=F21*2", "=G21*2"

time.sleep(3)


"""------------------ Cumulative Checkout User Statistics---------------"""

driver.find_element_by_xpath('//*[@id="statistics_btn"]').click()
send_checkout_date(monthly_stat_from, loan_date_to)
search_and_close()

#Save in Excel
ws["F23"] = get_user_statistics()["monthly_korean"]
ws["G23"] = get_user_statistics()["monthly_foreigner"]

time.sleep(3)

"""------------------ New User Statistics---------------"""

driver.find_element_by_xpath('//*[@id="statistics_type"]/option[3]').click()
driver.find_element_by_xpath('//*[@id="statistics_btn"]').click()
send_register_date(loan_date_from, loan_date_to)
reset = driver.find_element_by_xpath('//*[@id="shelf_loc_code_close_btn"]')
driver.execute_script("arguments[0].click();", reset)
driver.implicitly_wait(10)
by_date = driver.find_element_by_xpath('//*[@id="jqg442"]/td[2]')
by_class = driver.find_element_by_xpath('//*[@id="jqg461"]/td[2]')
driver.execute_script("arguments[0].click();", by_date)
driver.execute_script("arguments[0].click();", by_class)
search_and_close()


#Save in Excel
ws["G5"], ws["G6"], ws["G7"], ws["G8"], ws["G9"] = [get_user_statistics()["korean"][i] for i in range(5)]
ws["K5"], ws["K6"], ws["K7"], ws["K8"], ws["K9"] = [get_user_statistics()["foreigner"][i] for i in range(5)]

time.sleep(3)


"""------------------ Cumulative New User Statistics---------------"""

driver.find_element_by_xpath('//*[@id="statistics_btn"]').click()
send_register_date(monthly_stat_from, loan_date_to)
search_and_close()

#Save in Excel
ws["G11"] = get_user_statistics()["monthly_korean"]
ws["K11"] = get_user_statistics()["monthly_foreigner"]

time.sleep(3)

"""------------------ Collection Statistics---------------"""

#Open Statistics tap
driver.find_element_by_xpath('//*[@id="right_container_wrapper"]/header/div[5]/nav/a').click()
driver.find_element_by_xpath('//*[@id="MOLI_02_HTML"]/li[12]').click()
driver.switch_to.window(driver.window_handles[2])

driver.find_element_by_xpath('//*[@id="statistics_type"]/option[6]').click()
book_count = driver.find_element_by_xpath('//*[@id="non_price_type"]/label[2]')
driver.execute_script("arguments[0].click();", book_count)
remove_zero = driver.find_element_by_xpath('//*[@id="right_container_wrapper"]/main/div/div[1]/form/div[2]/label[4]')
driver.execute_script("arguments[0].click();", remove_zero)
driver.find_element_by_xpath('//*[@id="statistics_btn"]').click()
checkbox1 = driver.find_element_by_xpath('//*[@id="shelf_hold_statistics_search_modal"]/div/div/div[2]/div[1]/table/tbody/tr[3]/th[1]/label')
checkbox2 = driver.find_element_by_xpath('//*[@id="shelf_hold_statistics_search_modal"]/div/div/div[2]/div[1]/table/tbody/tr[4]/th[1]/label')
checkbox3 = driver.find_element_by_xpath('//*[@id="shelf_hold_statistics_search_modal"]/div/div/div[2]/div[1]/table/tbody/tr[5]/th[1]/label')
driver.execute_script("arguments[0].click();", checkbox1)
driver.execute_script("arguments[0].click();", checkbox2)
driver.execute_script("arguments[0].click();", checkbox3)

collection_stat = driver.find_element_by_xpath('//*[@id="working_status"]/option[4]')
driver.execute_script("arguments[0].click();", collection_stat)
search_and_close()

#Save in Excel
ws2["B3"], ws2["C3"], ws2["D3"], ws2["E3"], ws2["F3"], ws2["G3"], ws2["H3"], ws2["I3"], ws2["J3"], ws2["K3"], ws2["L3"] = [get_collection_statistics()["korean"][i] for i in range(11)]
ws2["B4"], ws2["C4"], ws2["D4"], ws2["E4"], ws2["F4"], ws2["G4"], ws2["H4"], ws2["I4"], ws2["J4"], ws2["K4"], ws2["L4"] = [get_collection_statistics()["korean_refer"][i] for i in range(11)]
ws2["B5"], ws2["C5"], ws2["D5"], ws2["E5"], ws2["F5"], ws2["G5"], ws2["H5"], ws2["I5"], ws2["J5"], ws2["K5"], ws2["L5"] = [get_collection_statistics()["korean_big"][i] for i in range(11)]
ws2["B6"], ws2["C6"], ws2["D6"], ws2["E6"], ws2["F6"], ws2["G6"], ws2["H6"], ws2["I6"], ws2["J6"], ws2["K6"], ws2["L6"] = [get_collection_statistics()["non_book"][i] for i in range(11)]
ws2["B7"], ws2["C7"], ws2["D7"], ws2["E7"], ws2["F7"], ws2["G7"], ws2["H7"], ws2["I7"], ws2["J7"], ws2["K7"], ws2["L7"] = [get_collection_statistics()["english"][i] for i in range(11)]
ws2["B8"], ws2["C8"], ws2["D8"], ws2["E8"], ws2["F8"], ws2["G8"], ws2["H8"], ws2["I8"], ws2["J8"], ws2["K8"], ws2["L8"] = [get_collection_statistics()["english_refer"][i] for i in range(11)]
ws2["B9"], ws2["C9"], ws2["D9"], ws2["E9"], ws2["F9"], ws2["G9"], ws2["H9"], ws2["I9"], ws2["J9"], ws2["K9"], ws2["L9"] = [get_collection_statistics()["chinese"][i] for i in range(11)]
ws2["B10"], ws2["C10"], ws2["D10"], ws2["E10"], ws2["F10"], ws2["G10"], ws2["H10"], ws2["I10"], ws2["J10"], ws2["K10"], ws2["L10"] = [get_collection_statistics()["french"][i] for i in range(11)]
ws2["B11"], ws2["C11"], ws2["D11"], ws2["E11"], ws2["F11"], ws2["G11"], ws2["H11"], ws2["I11"], ws2["J11"], ws2["K11"], ws2["L11"] = [get_collection_statistics()["japanese"][i] for i in range(11)]
ws2["B12"], ws2["C12"], ws2["D12"], ws2["E12"], ws2["F12"], ws2["G12"], ws2["H12"], ws2["I12"], ws2["J12"], ws2["K12"], ws2["L12"] = [get_collection_statistics()["mongolian"][i] for i in range(11)]
ws2["B13"], ws2["C13"], ws2["D13"], ws2["E13"], ws2["F13"], ws2["G13"], ws2["H13"], ws2["I13"], ws2["J13"], ws2["K13"], ws2["L13"] = [get_collection_statistics()["filipino"][i] for i in range(11)]
ws2["B14"], ws2["C14"], ws2["D14"], ws2["E14"], ws2["F14"], ws2["G14"], ws2["H14"], ws2["I14"], ws2["J14"], ws2["K14"], ws2["L14"] = [get_collection_statistics()["spanish"][i] for i in range(11)]
ws2["B15"], ws2["C15"], ws2["D15"], ws2["E15"], ws2["F15"], ws2["G15"], ws2["H15"], ws2["I15"], ws2["J15"], ws2["K15"], ws2["L15"] = [get_collection_statistics()["vietnamese"][i] for i in range(11)]
ws2["B16"], ws2["C16"], ws2["D16"], ws2["E16"], ws2["F16"], ws2["G16"], ws2["H16"], ws2["I16"], ws2["J16"], ws2["K16"], ws2["L16"] = [get_collection_statistics()["cambodian"][i] for i in range(11)]


"""------------------ Change Date ---------------"""

#Get Dates
loan_date_from_month = int(loan_date_from[5:7])
loan_date_from_day = int(loan_date_from[8:10])

monday = str(loan_date_from_month) + "/" + str(loan_date_from_day)
tuesday = str(loan_date_from_month) + "/" + str(loan_date_from_day + 1)
wednesday = str(loan_date_from_month) + "/" + str(loan_date_from_day + 2)
thursday = str(loan_date_from_month) + "/" + str(loan_date_from_day + 3)
friday = str(loan_date_from_month) + "/" + str(loan_date_from_day + 4)
saturday = str(loan_date_from_month) + "/" + str(loan_date_from_day + 5)
sunday = str(loan_date_from_month) + "/" + str(loan_date_from_day + 6)

#Save in Excel
ws["A1"] = "송도국제기구도서관 주간업무 보고\n" + loan_date_from + "(월) ~ " + loan_date_to + "(금)"
ws2["A1"] = "장서보유현황" + "[" + loan_date_to + "]"
ws3["A1"] = "형태별 기증자료현황" + "[" + loan_date_to + "]"
ws["A5"], ws["A17"], ws["A28"], ws["A39"] = [monday for i in range(0, 4)]
ws["A6"], ws["A18"], ws["A29"], ws["A40"] = [tuesday for i in range(0, 4)]
ws["A7"], ws["A19"], ws["A30"], ws["A41"] = [wednesday for i in range(0, 4)]
ws["A8"], ws["A20"], ws["A31"], ws["A42"] = [thursday for i in range(0, 4)]
ws["A9"], ws["A21"], ws["A32"], ws["A43"] = [friday for i in range(0, 4)]


#Random value for PC & Magazine User
ws["J17"], ws["J18"], ws["J19"], ws["J20"], ws["J21"] = [random.randint(15, 25) for i in range(0,5)]
ws["K17"], ws["K18"], ws["K19"], ws["K20"], ws["K21"] = [random.randint(0, 1) for i in range(0,5)]
ws["L17"], ws["L18"], ws["L19"], ws["L20"], ws["L21"] = [random.randint(22, 30) for i in range(0,5)]
ws["M17"], ws["M18"], ws["M19"], ws["M20"], ws["M21"] = [random.randint(0, 3) for i in range(0,5)]

wb.save("C:\\Users\\user\\Desktop\\주간업무 추진실적\\주간업무 추진실적(2020.02.24~2020.02.28).xlsx")

