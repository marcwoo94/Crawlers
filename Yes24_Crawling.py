import requests
import json
from collections import OrderedDict
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Color
from bs4 import BeautifulSoup


"""------------------------------------ Extract Link List ------------------------------------"""


def get_product_info(book):
    hrefs = book.find("a")["href"]
    link = "http://www.yes24.com" + hrefs
    return link


def get_page_info(page):
    result = requests.get("http://www.yes24.com/24/Category/Display/002001049005009?PageNumber=" + str(page))
    bs_obj = BeautifulSoup(result.content, "html.parser")
    shelf = bs_obj.find("div", {"class": "cCont_listArea"})
    books = shelf.findAll("div", {"class": "goods_name"})
    link_list = [get_product_info(book) for book in books]
    return link_list


result = []

for page in range(1, 6):
    result.extend(get_page_info(page))


def get_links():
    return result


"""---------------------------------Extract Book Information---------------------------------"""


def get_book_info(url):
    result = requests.get(url)
    bs_obj = BeautifulSoup(result.content, "html.parser")

    title = bs_obj.find("h2", {"class": "gd_name"}).text
    table_pub = bs_obj.find("span", {"class": "gd_pubArea"})
    contributors = table_pub.find("span", {"class": "gd_auth"})
    publisher = table_pub.find("span", {"class": "gd_pub"}).text
    table_price = bs_obj.find("div", {"class": "gd_infoTb"})
    price = table_price.findAll("em", {"class": "yes_m"})[0].text
    table_info = bs_obj.find("table", {"class": "tb_detail01"})
    date = table_info.findAll("td", {"class": "cell_2col"})[0].text
    category_1 = bs_obj.find("div", {"class": "basicListType communtyHide"})
    category_2 = category_1.findAll("a")
    category_3 = [item.text for item in category_2]
    category_4 = str(list(OrderedDict.fromkeys(category_3)))
    category_5 = category_4.replace("[", "").replace("]", "").replace(",", " |").replace("'", "")

    try:
        author = contributors.findAll("a")[0].text
    except:
        try:
            author = contributors.text.split("\r\n")[1].split("        ")[1]
        except:
            author = ""

    try:
        cover = bs_obj.find("span", {"class": "gd_feature"}).text.split(" ")[1].split(",")[0]
    except:
        try:
            cover = bs_obj.find("span", {"class": "gd_feature"}).text.split(" ")[1]
        except:
            cover = ""

    try:
        translator = contributors.findAll("a")[1].text
    except:
        translator = ""

    try:
        isbn = table_info.findAll("td", {"class": "cell_2col"})[2].text
    except:
        isbn = ""


    dictionary1 = {}
    dictionary1["title"] = title
    dictionary1["author"] = author
    dictionary1["translator"] = translator
    dictionary1["publisher"] = publisher
    dictionary1["price"] = price
    dictionary1["cover"] = cover
    dictionary1["ISBN"] = isbn
    dictionary1["date"] = date
    dictionary1["category"] = category_5

    return dictionary1


urls = get_links()


def get_info():
    book_info = [get_book_info(url) for url in urls]
    return book_info


"""--------------------------------------Save in Excel--------------------------------------"""


j1 = get_info()
j2 = json.dumps(j1)
j3 = json.loads(j2)

wb = Workbook()
ws1 = wb.active
ws1.freeze_panes = "A2"

ws1["A1"] = "제목"
ws1["A1"].font = Font(size=12, bold=True)
ws1["A1"].fill = PatternFill(patternType="solid", fgColor=Color("B8CCE4"))
ws1["B1"] = "저자"
ws1["B1"].font = Font(size=12, bold=True)
ws1["B1"].fill = PatternFill(patternType="solid", fgColor=Color("B8CCE4"))
ws1["C1"] = "역자"
ws1["C1"].font = Font(size=12, bold=True)
ws1["C1"].fill = PatternFill(patternType="solid", fgColor=Color("B8CCE4"))
ws1["D1"] = "출판사"
ws1["D1"].font = Font(size=12, bold=True)
ws1["D1"].fill = PatternFill(patternType="solid", fgColor=Color("B8CCE4"))
ws1["E1"] = "가격"
ws1["E1"].font = Font(size=12, bold=True)
ws1["E1"].fill = PatternFill(patternType="solid", fgColor=Color("B8CCE4"))
ws1["F1"] = "표지"
ws1["F1"].font = Font(size=12, bold=True)
ws1["F1"].fill = PatternFill(patternType="solid", fgColor=Color("B8CCE4"))
ws1["G1"] = "ISBN"
ws1["G1"].font = Font(size=12, bold=True)
ws1["G1"].fill = PatternFill(patternType="solid", fgColor=Color("B8CCE4"))
ws1["H1"] = "출판일"
ws1["H1"].font = Font(size=12, bold=True)
ws1["H1"].fill = PatternFill(patternType="solid", fgColor=Color("B8CCE4"))
ws1["I1"] = "카테고리"
ws1["I1"].font = Font(size=12, bold=True)
ws1["I1"].fill = PatternFill(patternType="solid", fgColor=Color("B8CCE4"))


for row_index in range(1, len(j1)+1):
    ws1.cell(row=row_index + 1, column=1).value = j3[row_index - 1]["title"]
    ws1.cell(row=row_index + 1, column=2).value = j3[row_index - 1]["author"]
    ws1.cell(row=row_index + 1, column=3).value = j3[row_index - 1]["translator"]
    ws1.cell(row=row_index + 1, column=4).value = j3[row_index - 1]["publisher"]
    ws1.cell(row=row_index + 1, column=5).value = j3[row_index - 1]["price"]
    ws1.cell(row=row_index + 1, column=6).value = j3[row_index - 1]["cover"]
    ws1.cell(row=row_index + 1, column=7).value = j3[row_index - 1]["ISBN"]
    ws1.cell(row=row_index + 1, column=8).value = j3[row_index - 1]["date"]
    ws1.cell(row=row_index + 1, column=9).value = j3[row_index - 1]["category"]


wb.save("C:\\Users\\user\\Desktop\\우제현\\Yes24_Top100.xlsx")
