import requests
import json
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Color, Alignment, Border, Side
from bs4 import BeautifulSoup
from selenium import webdriver
import os


"""------------------------------------ Log In to Yes24 --------------------------------------"""

driver = webdriver.Chrome(os.path.dirname(os.path.abspath(__file__)) + "\chromedriver.exe")
driver.implicitly_wait(3)
driver.get("https://www.yes24.com/Templates/FTLogin.aspx")
driver.maximize_window()
driver.find_element_by_name("SMemberID").send_keys("*******")
driver.find_element_by_name("SMemberPassword").send_keys("*******")
driver.find_element_by_xpath("//*[@id='btnLogin']/span/em").click()


"""------------------------------------ Extract Link List ------------------------------------"""

def get_product_info(book):
    link = book.find("a", class_="pd_a")["href"]
    return link

def get_page_info():
    driver.get("http://ssl.yes24.com/Cart/Cart")
    result = driver.page_source
    bs_obj = BeautifulSoup(result, "html.parser")
    books = bs_obj.find_all("td", class_="le")
    link_list = [get_product_info(book) for book in books]
    return link_list

result = []

def get_links():
    result.extend(get_page_info())
    return result


"""---------------------------------Extract Book Information---------------------------------"""


def get_book_info(url):
    result = requests.get(url)
    bs_obj = BeautifulSoup(result.content, "html.parser")

    # 변수 지정
    gd_pub = bs_obj.find("span", class_="gd_pub")
    table_price = bs_obj.find("div", class_="gd_infoTb")
    original_price_tag = bs_obj.select("div.gd_infoTb th")[0].text.strip()
    contributors = bs_obj.find("span", class_="gd_auth")
    category_table = bs_obj.find("dl", class_="yesAlertDl")
    category_list = category_table.find_all("li")


    try: title = bs_obj.find("h2", class_="gd_name").text  # 제목
    except: title = ""
    try: isbn = bs_obj.find_all("td", class_="txt lastCol")[2].text.strip() #ISBN
    except: isbn = ""
    try: publisher = gd_pub.find("a").text  # 출판사
    except: publisher = ""
    try: date = bs_obj.find_all("td", class_="txt lastCol")[0].text.replace("년 ","/").replace("월 ","/").replace("일","")  # 출판일
    except: date = ""
    if original_price_tag == "정가":  #가격
       price = table_price.find_all("em", class_="yes_m")[0].text.split("원")[0].replace(",", "")
    else:
        price = bs_obj.select("div.gd_infoTb tr td span.nor_price em")[0].text.split("원")[0].replace(",", "")
    try: author = contributors.find_all("a")[0].text  # 저자
    except:
        try: author = contributors.text.strip()
        except: author = ""
    try: translator = contributors.find_all("a")[1].text  # 역자
    except: translator = ""
    try: remarks = bs_obj.find("span", class_="gd_feature").text.split("[")[1].split("]")[0].strip()  # 표지
    except: remarks = ""
    try: category = category_list[0].find_all("a")[1].text + ", " + category_list[1].find_all("a")[1].text + ", " + category_list[2].find_all("a")[1].text  #카테고리
    except:
        try: category = category_list[0].find_all("a")[1].text + ", " + category_list[1].find_all("a")[1].text
        except: category = category_list[0].find_all("a")[1].text


    dictionary1 = {}
    dictionary1["title"] = title
    dictionary1["author"] = author
    dictionary1["translator"] = translator
    dictionary1["publisher"] = publisher
    dictionary1["price"] = int(price)
    dictionary1["remarks"] = remarks
    dictionary1["date"] = date
    dictionary1["category"] = category
    try:
        dictionary1["ISBN"] = int(isbn)
    except:
        dictionary1["ISBN"] = "ISBN 확인"

    return dictionary1

urls = get_links()

def get_info():
    book_info = [get_book_info(url) for url in urls]
    return book_info


"""--------------------------------------Save in Excel--------------------------------------"""


def save_in_excel():

    # dictionary 덤프
    j1 = get_info()
    j2 = json.dumps(j1)
    j3 = json.loads(j2)

    # Excel 파일 생성
    wb = Workbook()
    ws1 = wb.active
    ws1.title = "총괄표"
    ws2 = wb.create_sheet("구입도서목록", 1)


    # 서식 설정
    align_center = Alignment(horizontal="center", vertical="center")
    align_left = Alignment(horizontal="left", vertical="center")
    thin_border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))
    title_font_1 = Font(name="맑은 고딕", size=18, bold=True)
    title_font_2 = Font(name="맑은 고딕", size=14, bold=True)
    title_font_3 = Font(name="맑은 고딕", size=14, bold=False)
    title_font_4 = Font(name="맑은 고딕", size=11, bold=True)
    title_fill = PatternFill(patternType="solid", fgColor=Color("FDE9D9"))
    yellow_fill = PatternFill(patternType="solid", fgColor=Color("FFFF99"))
    red_fill = PatternFill(patternType="solid", fgColor=Color("FF0000"))
    white_fill = PatternFill(patternType="solid", fgColor=Color("ffffff"))
    currency_format = "₩#,###"
    percentage_format = "#%"
    number_format = "#############"


    # ws2 행 넓이, 열 높이 설정
    ws2.row_dimensions[1].height = 40
    ws2.column_dimensions['A'].width = 5
    ws2.column_dimensions['B'].width = 25
    ws2.column_dimensions['C'].width = 45
    ws2.column_dimensions['D'].width = 25
    ws2.column_dimensions['E'].width = 25
    ws2.column_dimensions['F'].width = 15
    ws2.column_dimensions['G'].width = 10
    ws2.column_dimensions['H'].width = 15
    ws2.column_dimensions['I'].width = 15
    ws2.column_dimensions['J'].width = 15
    ws2.column_dimensions['K'].width = 30


    # ws2 행 이름 설정
    ws2.merge_cells("A1:I1")
    ws2["A1"] = "송도국제기구도서관 2020년 자료구입 선정목록"
    ws2["A1"].font = title_font_2
    ws2["A1"].alignment = align_center
    column_title_2 = ["번호", "주제명", "서명", "저자", "발행자", "발행년", "책 수", "단가", "구입가격", "ISBN", "비고"]
    for col in range(0, 11):
        ws2.cell(row=2, column=col + 1).value = column_title_2[col]
        ws2.cell(row=2, column=col + 1).fill = title_fill
        ws2.cell(row=2, column=col + 1).alignment = align_center

    # ws2에 dictionary 엑셀에 입력
    for row_index in range(1, len(j1)+1):
        ws2.cell(row=row_index + 3, column=1).value = row_index
        ws2.cell(row=row_index + 3, column=2).value = j3[row_index - 1]["category"]
        ws2.cell(row=row_index + 3, column=3).value = j3[row_index - 1]["title"]
        ws2.cell(row=row_index + 3, column=4).value = j3[row_index - 1]["author"]
        ws2.cell(row=row_index + 3, column=5).value = j3[row_index - 1]["publisher"]
        ws2.cell(row=row_index + 3, column=6).value = j3[row_index - 1]["date"]
        ws2.cell(row=row_index + 3, column=8).value = f"={j3[row_index - 1]['price']}/G{row_index + 3}"
        ws2.cell(row=row_index + 3, column=9).value = f"=G{row_index + 3}*H{row_index + 3}"
        ws2.cell(row=row_index + 3, column=10).value = j3[row_index - 1]["ISBN"]
        ws2.cell(row=row_index + 3, column=11).value = j3[row_index - 1]["remarks"]

        contain_set = "세트"
        contain_plus = " + "
        find_set = j3[row_index - 1]["title"].find(contain_set)
        find_plus = j3[row_index - 1]["title"].find(contain_plus)
        if find_set == -1 and find_plus == -1:
            ws2.cell(row=row_index + 3, column=7).value = 1
        else:
            ws2.cell(row=row_index + 3, column=7).value = "권 수 확인"



    # ws2 테두리 설정
    for row in ws2["A2":f"K{len(j1)+3}"]:
        for cell in row:
            cell.border = thin_border

    # ws2 ISBN, 비고, 세트 색 설정
    for row in ws2["J2":f"K{len(j1)+3}"]:
        for cell in row:
            cell.fill = yellow_fill
    for row in ws2["B4":f"B{len(j1)+3}"]:
        for cell in row:
            cell.fill = yellow_fill
    for row in ws2["G4":f"G{len(j1)+3}"]:
        for cell in row:
            if cell.value == "권 수 확인":
                cell.fill = red_fill
            else:
                cell.fill = white_fill
    for row in ws2["J4":f"J{len(j1)+3}"]:
        for cell in row:
            if cell.value == "ISBN 확인":
                cell.fill = red_fill
            else:
                cell.fill = yellow_fill

    # ws2 왼쪽 정렬
    for row in ws2["A3":f"K{len(j1)+3}"]:
        for cell in row:
            cell.alignment = align_left

    # ws2 통화(₩) 설정
    for row in ws2["H3":f"I{len(j1)+3}"]:
        for cell in row:
            cell.number_format = currency_format

    # ws2 ISBN 숫자 설정
    for row in ws2["J4":f"J{len(j1)+3}"]:
        for cell in row:
            cell.number_format = number_format


    # ws1 행 이름 설정
    ws1.merge_cells("A1:F1")
    ws1["A1"] = "송도국제기구도서관 2020년 자료구입 총괄표"
    ws1["A1"].font = title_font_1
    ws1["A1"].alignment = align_center
    column_title_1 = ["구분", "종수", "권수", "비율(권수)", "금액(원)", "비율(금액)"]
    for col in range(0, 6):
        ws1.cell(row=2, column=col + 1).value = column_title_1[col]
        ws1.cell(row=2, column=col + 1).fill = title_fill

    # ws1 열 이름 설정
    row_title = ["합계", "총류", "철학", "종교", "사회과학", "자연과학", "기술과학", "예술", "언어", "문학", "역사"]
    for row in range(0, 11):
        ws1.cell(row= row + 3, column=1).value = row_title[row]

    # ws1 행 넓이, 열 높이 설정
    abc = ["A", "B", "C", "D", "E", "F", "G", "H", "I"]
    ws1.row_dimensions[1].height = 40

    for i in range(2, 14):
        ws1.row_dimensions[i].height = 30

    for i in range(0, 6):
        ws1.column_dimensions[f"{abc[i]}"].width = 15


    # 합계 함수 입력
    for i in range(0, 5):
        ws1[f"{abc[i+1]}3"] = f"=SUM({abc[i+1]}4: {abc[i+1]}13)"

    # IF 함수 입력
    for i in range(4, 14):
        ws1[f"B{i}"] = f'=COUNTIF(구입도서목록!B4:B{len(j1)+3},"{row_title[i-3]}")'
        ws1[f"C{i}"] = f'=SUMIF(구입도서목록!B4:B{len(j1)+3}, "{row_title[i-3]}", 구입도서목록!G4:G{len(j1)+3})'
        ws1[f"D{i}"] = f'=C{i}/C3'
        ws1[f"E{i}"] = f'=SUMIF(구입도서목록!B4:B{len(j1)+3}, "{row_title[i-3]}", 구입도서목록!I4:I{len(j1)+3})'
        ws1[f"F{i}"] = f'=E{i}/E3'


    # ws1 테두리 설정, 가운데 정렬, 폰트 설정
    for row in ws1["A2":"F13"]:
        for cell in row:
            cell.border = thin_border
            cell.alignment = align_center
            cell.font = title_font_3

    for row in ws1["A2":"F3"]:
        for cell in row:
            cell.font = title_font_2

    # ws1 통화(₩) 설정
    for row in ws1["E3":"E13"]:
        for cell in row:
            cell.number_format = currency_format

    # ws1 퍼센트(%) 설정
    for row in ws1["D3":"D13"]:
        for cell in row:
            cell.number_format = percentage_format

    for row in ws1["F3": "F13"]:
        for cell in row:
            cell.number_format = percentage_format

    #기타 설정
    ws2["F3"] = f'=COUNT(A4:A{len(j1)+3})&"종"'
    ws2["G3"] = f'=SUM(G4:G{len(j1)+3})&"권"'
    ws2["I3"] = f'=SUM(I4:I{len(j1)+3})'

    ws2["F3"].font = title_font_4
    ws2["G3"].font = title_font_4
    ws2["I3"].font = title_font_4
    ws2["I3"].number_format = currency_format


    wb.save(os.path.dirname(os.path.abspath(__file__)) + "\구입자료 목록.xlsx")

save_in_excel()
