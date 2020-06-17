from PyQt5.QtWidgets import *
from PyQt5 import QtCore, QtGui, QtWidgets
from selenium import webdriver
from bs4 import BeautifulSoup
from openpyxl import load_workbook
from datetime import date, timedelta
import time
import os

class Ui_Alpas(object):
    def setupUi(self, Alpas):
        Alpas.setObjectName("Alpas")
        Alpas.resize(530, 300)
        self.stat_from = QtWidgets.QDateEdit(Alpas)
        self.stat_from.setDate(QtCore.QDate.currentDate())
        self.stat_from.setGeometry(QtCore.QRect(290, 80, 200, 30))
        font = QtGui.QFont()
        font.setFamily("-윤고딕340")
        font.setPointSize(12)
        self.stat_from.setFont(font)
        self.stat_from.setCalendarPopup(True)
        self.stat_from.setObjectName("stat_from")
        self.alpas_id = QtWidgets.QLineEdit(Alpas)
        self.alpas_id.setGeometry(QtCore.QRect(40, 80, 200, 30))
        font = QtGui.QFont()
        font.setFamily("-윤고딕340")
        font.setPointSize(12)
        self.alpas_id.setFont(font)
        self.alpas_id.setObjectName("alpas_id")
        self.alpas_pw = QtWidgets.QLineEdit(Alpas)
        self.alpas_pw.setGeometry(QtCore.QRect(40, 125, 200, 30))
        font = QtGui.QFont()
        font.setFamily("-윤고딕340")
        font.setPointSize(12)
        self.alpas_pw.setFont(font)
        self.alpas_pw.setObjectName("alpas_pw")
        self.stat_date = QtWidgets.QLabel(Alpas)
        self.stat_date.setGeometry(QtCore.QRect(290, 40, 89, 28))
        font = QtGui.QFont()
        font.setFamily("Noto Sans KR Bold")
        font.setPointSize(14)
        font.setBold(True)
        font.setWeight(75)
        self.stat_date.setFont(font)
        self.stat_date.setObjectName("stat_date")
        self.alpas = QtWidgets.QLabel(Alpas)
        self.alpas.setGeometry(QtCore.QRect(40, 40, 51, 22))
        font = QtGui.QFont()
        font.setFamily("Noto Sans KR Bold")
        font.setPointSize(14)
        font.setBold(True)
        font.setWeight(75)
        self.alpas.setFont(font)
        self.alpas.setObjectName("alpas")
        self.pushButton = QtWidgets.QPushButton(Alpas)
        self.pushButton.setGeometry(QtCore.QRect(290, 190, 200, 68))
        font = QtGui.QFont()
        font.setFamily("Noto Sans KR Bold")
        font.setPointSize(12)
        font.setBold(True)
        font.setWeight(75)
        self.pushButton.setFont(font)
        self.pushButton.setObjectName("pushButton")
        self.file_name = QtWidgets.QLabel(Alpas)
        self.file_name.setGeometry(QtCore.QRect(40, 190, 72, 28))
        font = QtGui.QFont()
        font.setFamily("Noto Sans KR Bold")
        font.setPointSize(14)
        font.setBold(True)
        font.setWeight(75)
        self.file_name.setFont(font)
        self.file_name.setObjectName("file_name")
        self.file_name_original = QtWidgets.QLineEdit(Alpas)
        self.file_name_original.setGeometry(QtCore.QRect(40, 228, 200, 30))
        font = QtGui.QFont()
        font.setFamily("Noto Sans KR Medium")
        font.setPointSize(12)
        self.file_name_original.setFont(font)
        self.file_name_original.setObjectName("file_name_original")

        self.retranslateUi(Alpas)
        QtCore.QMetaObject.connectSlotsByName(Alpas)

        self.alpas_pw.setEchoMode(QLineEdit.Password)
        self.pushButton.clicked.connect(self.get_statistics)

    def retranslateUi(self, Alpas):
        _translate = QtCore.QCoreApplication.translate
        Alpas.setWindowTitle(_translate("Alpas", "Alpas_Statistics"))
        self.stat_from.setDisplayFormat(_translate("Alpas", "yyyy/MM/dd"))
        self.alpas_id.setPlaceholderText(_translate("Alpas", "ID"))
        self.alpas_pw.setPlaceholderText(_translate("Alpas", "PW"))
        self.stat_date.setText(_translate("Alpas", "통계 시작일"))
        self.alpas.setText(_translate("Alpas", "알파스"))
        self.pushButton.setText(_translate("Alpas", "주간실적 저장하기"))
        self.file_name.setText(_translate("Alpas", "파일 이름"))
        self.file_name_original.setPlaceholderText(_translate("Alpas", "지난 주 통계 파일 이름"))

    def get_statistics(self):

        def send_checkout_date(from_, to_):
            driver.implicitly_wait(10)
            driver.find_element_by_id("loan_date_from").clear()
            driver.find_element_by_id("loan_date_from").send_keys(from_)
            driver.find_element_by_id("loan_date_to").clear()
            driver.find_element_by_id("loan_date_to").send_keys(to_)
            driver.implicitly_wait(5)

        def send_return_date(from_, to_):
            driver.implicitly_wait(10)
            driver.find_element_by_id("return_date_from").clear()
            driver.find_element_by_id("return_date_from").send_keys(from_)
            driver.find_element_by_id("return_date_to").clear()
            driver.find_element_by_id("return_date_to").send_keys(to_)
            time.sleep(1)

        def send_register_date(from_, to_):
            driver.implicitly_wait(10)
            driver.find_element_by_id("reg_date_from").clear()
            driver.find_element_by_id("reg_date_from").send_keys(from_)
            driver.find_element_by_id("reg_date_to").clear()
            driver.find_element_by_id("reg_date_to").send_keys(to_)
            time.sleep(1)

        def send_shelf_date(from_, to_):
            driver.implicitly_wait(10)
            driver.find_element_by_id("shelf_date_from").clear()
            driver.find_element_by_id("shelf_date_from").send_keys(from_)
            driver.find_element_by_id("shelf_date_to").clear()
            driver.find_element_by_id("shelf_date_to").send_keys(to_)
            time.sleep(1)

        def search_and_close():
            driver.implicitly_wait(10)
            try:
                search = driver.find_element_by_xpath('//*[@id="statistics_search_btn"]')
            except:
                search = driver.find_element_by_xpath('//*[@id="shelf_hold_statistics_search_modal"]/div/div/div[3]/div/div[2]/button[1]')
            click(search)
            driver.implicitly_wait(120)
            msg_btn_1 = driver.find_element_by_xpath('//*[@id="msg_btn_1"]')
            click(msg_btn_1)
            driver.implicitly_wait(10)
            close_button = driver.find_element_by_xpath('//*[@id="pusrchaesDataPrintdeSearchClose"]')
            click(close_button)
            time.sleep(1)

        def get_page_info():
            result = driver.page_source
            bs_obj = BeautifulSoup(result, "html.parser")
            table = bs_obj.find("table", class_="display compact cell-border table table-bordered ui-jqgrid-btable ui-common-table")
            tr = table.find_all("tr")

            general_works = tr[1].find_all("td")
            philosophy = tr[2].find_all("td")
            religion = tr[3].find_all("td")
            social_sciences = tr[4].find_all("td")
            natural_sciences = tr[5].find_all("td")
            technology = tr[6].find_all("td")
            arts = tr[7].find_all("td")
            language = tr[8].find_all("td")
            literature = tr[9].find_all("td")
            history = tr[10].find_all("td")

            dictionary = {}
            dictionary["general_works"] = [int(general_works[i].text) for i in range(4, 9)]
            dictionary["philosophy"] = [int(philosophy[i].text) for i in range(4, 9)]
            dictionary["religion"] = [int(religion[i].text) for i in range(4, 9)]
            dictionary["social_sciences"] = [int(social_sciences[i].text) for i in range(4, 9)]
            dictionary["natural_sciences"] = [int(natural_sciences[i].text) for i in range(4, 9)]
            dictionary["technology"] = [int(technology[i].text) for i in range(4, 9)]
            dictionary["arts"] = [int(arts[i].text) for i in range(4, 9)]
            dictionary["language"] = [int(language[i].text) for i in range(4, 9)]
            dictionary["literature"] = [int(literature[i].text) for i in range(4, 9)]
            dictionary["history"] = [int(history[i].text) for i in range(4, 9)]

            dictionary["monthly_general_works"] = int(general_works[3].text)
            dictionary["monthly_philosophy"] = int(philosophy[3].text)
            dictionary["monthly_religion"] = int(religion[3].text)
            dictionary["monthly_social_sciences"] = int(social_sciences[3].text)
            dictionary["monthly_natural_sciences"] = int(natural_sciences[3].text)
            dictionary["monthly_technology"] = int(technology[3].text)
            dictionary["monthly_arts"] = int(arts[3].text)
            dictionary["monthly_language"] = int(language[3].text)
            dictionary["monthly_literature"] = int(literature[3].text)
            dictionary["monthly_history"] = int(history[3].text)

            return dictionary

        def get_user_statistics():
            result = driver.page_source
            bs_obj = BeautifulSoup(result, "html.parser")
            table = bs_obj.find("table", class_="display compact cell-border table table-bordered ui-jqgrid-btable ui-common-table")
            tr = table.find_all("tr")

            male = tr[1].find_all("td")
            female = tr[2].find_all("td")
            male_student = tr[3].find_all("td")
            female_student = tr[4].find_all("td")
            male_youth = tr[5].find_all("td")
            female_youth = tr[6].find_all("td")
            employee = tr[7].find_all("td")
            digital = tr[8].find_all("td")
            good_member = tr[9].find_all("td")
            foreigner = tr[10].find_all("td")
            good_member_chungla = tr[11].find_all("td")
            good_member_heavy_reader = tr[12].find_all("td")
            related_organization = tr[13].find_all("td")
            reference_room = tr[14].find_all("td")
            vision_impaired = tr[15].find_all("td")
            hearing_impaired = tr[16].find_all("td")
            physical_disabled = tr[17].find_all("td")
            brain_disabled = tr[18].find_all("td")
            traveling_library = tr[19].find_all("td")
            renal_disorder = tr[20].find_all("td")
            mental_disorder = tr[21].find_all("td")
            other_disorder = tr[22].find_all("td")
            honorary_member = tr[23].find_all("td")
            do_not_delete = tr[24].find_all("td")

            dictionary = {}
            dictionary["korean"] = [int(male[i].text) + int(female[i].text) + int(male_student[i].text) +
                                    int(female_student[i].text) + int(male_youth[i].text) + int(female_youth[i].text) +
                                    int(employee[i].text) + int(digital[i].text) + int(good_member[i].text) +
                                    int(good_member_chungla[i].text) + int(good_member_heavy_reader[i].text) +
                                    int(related_organization[i].text) + int(reference_room[i].text) + int(vision_impaired[i].text) +
                                    int(hearing_impaired[i].text) + int(physical_disabled[i].text) + int(brain_disabled[i].text) +
                                    int(traveling_library[i].text) + int(renal_disorder[i].text) + int(mental_disorder[i].text) +
                                    int(other_disorder[i].text) + int(honorary_member[i].text) + int(do_not_delete[i].text) for i in range(4, 9)]

            dictionary["foreigner"] = [int(foreigner[i].text) for i in range(4, 9)]

            dictionary["monthly_korean"] = int(male[3].text) + int(female[3].text) + int(male_student[3].text) + int(female_student[3].text)\
                                           + int(male_youth[3].text) + int(female_youth[3].text) + int(employee[3].text) + int(digital[3].text)\
                                           + int(good_member[3].text) + int(good_member_chungla[3].text) + int(good_member_heavy_reader[3].text)\
                                           + int(related_organization[3].text) + int(reference_room[3].text) + int(vision_impaired[3].text)\
                                           + int(hearing_impaired[3].text) + int(physical_disabled[3].text) + int(brain_disabled[3].text)\
                                           + int(traveling_library[3].text) + int(renal_disorder[3].text) + int(mental_disorder[3].text)\
                                           + int(other_disorder[3].text) + int(honorary_member[3].text) + int(do_not_delete[3].text)

            dictionary["monthly_foreigner"] = int(foreigner[3].text)

            return dictionary

        def get_collection_statistics():
            result = driver.page_source
            bs_obj = BeautifulSoup(result, "html.parser")
            table = bs_obj.find("table", class_="display compact cell-border table table-bordered ui-jqgrid-btable ui-common-table")
            tr = table.find_all("tr")

            korean_big = tr[1].find_all("td")
            korean = tr[2].find_all("td")
            korean_refer = tr[3].find_all("td")
            non_book = tr[4].find_all("td")
            chinese = tr[5].find_all("td")
            french = tr[6].find_all("td")
            english = tr[7].find_all("td")
            japanese = tr[8].find_all("td")
            mongolian = tr[9].find_all("td")
            filipino = tr[10].find_all("td")
            english_refer = tr[11].find_all("td")
            spanish = tr[12].find_all("td")
            vietnamese = tr[13].find_all("td")
            cambodian = tr[14].find_all("td")

            sum = tr[-1].find_all("td")


            dictionary = {}
            dictionary["korean_big"] = [int(korean_big[i].text) for i in range(3, 14)]
            dictionary["korean"] = [int(korean[i].text) for i in range(3, 14)]
            dictionary["korean_refer"] = [int(korean_refer[i].text) for i in range(3, 14)]
            dictionary["non_book"] = [int(non_book[i].text) for i in range(3, 14)]
            dictionary["chinese"] = [int(chinese[i].text) for i in range(3, 14)]
            dictionary["french"] = [int(french[i].text) for i in range(3, 14)]
            dictionary["english"] = [int(english[i].text) for i in range(3, 14)]
            dictionary["japanese"] = [int(japanese[i].text) for i in range(3, 14)]
            dictionary["mongolian"] = [int(mongolian[i].text) for i in range(3, 14)]
            dictionary["filipino"] = [int(filipino[i].text) for i in range(3, 14)]
            dictionary["english_refer"] = [int(english_refer[i].text) for i in range(3, 14)]
            dictionary["spanish"] = [int(spanish[i].text) for i in range(3, 14)]
            dictionary["vietnamese"] = [int(vietnamese[i].text) for i in range(3, 14)]
            dictionary["cambodian"] = [int(cambodian[i].text) for i in range(3, 14)]

            dictionary["sum"] = [int(sum[i].text) for i in range(3, 14)]

            return dictionary

        def get_purchase_statistics():
            result = driver.page_source
            bs_obj = BeautifulSoup(result, "html.parser")
            table = bs_obj.find("table",
                                class_="display compact cell-border table table-bordered ui-jqgrid-btable ui-common-table")
            tr = table.find_all("tr")

            sum = tr[-1].find_all("td")

            dictionary = {}
            dictionary["sum"] = [int(sum[i].text) for i in range(3, 14)]

            return dictionary
            time.sleep(1)

        def get_donation_statistics():
            result = driver.page_source
            bs_obj = BeautifulSoup(result, "html.parser")
            table = bs_obj.find("table",
                                class_="display compact cell-border table table-bordered ui-jqgrid-btable ui-common-table")
            tr = table.find_all("tr")

            sum = tr[-1].find_all("td")

            dictionary = {}
            dictionary["sum"] = [int(sum[i].text) for i in range(3, 14)]

            return dictionary
            time.sleep(1)

        def click(x):
            driver.execute_script("arguments[0].click();", x)
            time.sleep(0.5)

        loan_date_input = self.stat_from.text()
        year = loan_date_input.split("/")[0]
        month = loan_date_input.split("/")[1]
        day = loan_date_input.split("/")[2]
        loan_date_from_dt = date(int(year), int(month), int(day))
        loan_date_to_dt = loan_date_from_dt + timedelta(days=4)
        loan_date_from = str(loan_date_from_dt).replace("-", "/")
        loan_date_to = str(loan_date_to_dt).replace("-", "/")
        monthly_stat_from = (str(loan_date_from_dt)[0:8] + "01").replace("-", "/")

        wb = load_workbook(os.path.dirname(os.path.abspath(__file__)) + f"\{self.file_name_original.text()}.xlsx")
        ws = wb["1.이용실적(日)"]
        ws3 = wb["3.장서확충실적"]
        ws4 = wb["4.장서보유현황"]

        # Login
        driver = webdriver.Chrome(os.path.dirname(os.path.abspath(__file__)) + "\chromedriver.exe")
        driver.implicitly_wait(3)
        driver.get("http://152.99.43.46:28180/METIS/")
        driver.maximize_window()
        driver.find_element_by_name("main_login_user_id").send_keys(self.alpas_id.text())
        driver.find_element_by_id("user_pw").send_keys(self.alpas_pw.text())
        login = driver.find_element_by_xpath("/html/body/div[2]/div/button")
        click(login)
        time.sleep(3)

        # Close Pop up
        close_popup = driver.find_element_by_xpath('//*[@id="closeNoticePopup"]')
        click(close_popup)

        # Open Statistics tap
        menu = driver.find_element_by_xpath('//*[@id="right_container_wrapper"]/header/div[5]/nav/a')
        click(menu)
        statistics = driver.find_element_by_xpath('//*[@id="COLI_01_HTML"]/li[13]')
        click(statistics)
        driver.switch_to.window(driver.window_handles[1])

        # Get Check Out Statistics
        driver.find_element_by_xpath('//*[@id="statistics_type"]/option[6]').click()  # Choose Checkout Statistics
        show_zero = driver.find_element_by_xpath('//*[@id="right_container_wrapper"]/main/div/div[1]/form/div[2]/label[4]')
        click(show_zero)
        stat_btn = driver.find_element_by_xpath('//*[@id="statistics_btn"]')  # Click Statistics button
        click(stat_btn)
        time.sleep(1)
        send_checkout_date(loan_date_from, loan_date_to)  # Enter checkout date
        driver.find_element_by_xpath('//*[@id="shelf_loc_code_list"]/option[36]').click()  # Choose Songdo Library of International Organization
        add_songdo = driver.find_element_by_xpath('//*[@id="shelf_loc_code_btn"]')
        click(add_songdo)
        by_date = driver.find_element_by_xpath('//*[@id="jqg82"]/td[2]')  # Adjust matrix
        click(by_date)

        try:
            search_and_close()
            get_page_info()  # Parsing

            # Save in Excel
            ws["F35"], ws["F36"], ws["F37"], ws["F38"], ws["F39"] = [get_page_info()["general_works"][i] for i in range(5)]
            ws["G35"], ws["G36"], ws["G37"], ws["G38"], ws["G39"] = [get_page_info()["philosophy"][i] for i in range(5)]
            ws["H35"], ws["H36"], ws["H37"], ws["H38"], ws["H39"] = [get_page_info()["religion"][i] for i in range(5)]
            ws["I35"], ws["I36"], ws["I37"], ws["I38"], ws["I39"] = [get_page_info()["social_sciences"][i] for i in range(5)]
            ws["J35"], ws["J36"], ws["J37"], ws["J38"], ws["J39"] = [get_page_info()["natural_sciences"][i] for i in range(5)]
            ws["K35"], ws["K36"], ws["K37"], ws["K38"], ws["K39"] = [get_page_info()["technology"][i] for i in range(5)]
            ws["L35"], ws["L36"], ws["L37"], ws["L38"], ws["L39"] = [get_page_info()["arts"][i] for i in range(5)]
            ws["M35"], ws["M36"], ws["M37"], ws["M38"], ws["M39"] = [get_page_info()["language"][i] for i in range(5)]
            ws["N35"], ws["N36"], ws["N37"], ws["N38"], ws["N39"] = [get_page_info()["literature"][i] for i in range(5)]
            ws["O35"], ws["O36"], ws["O37"], ws["O38"], ws["O39"] = [get_page_info()["history"][i] for i in range(5)]

        except IndexError:
            # Save in Excel
            ws["F35"], ws["F36"], ws["F37"], ws["F38"], ws["F39"] = [0 for i in range(5)]
            ws["G35"], ws["G36"], ws["G37"], ws["G38"], ws["G39"] = [0 for i in range(5)]
            ws["H35"], ws["H36"], ws["H37"], ws["H38"], ws["H39"] = [0 for i in range(5)]
            ws["I35"], ws["I36"], ws["I37"], ws["I38"], ws["I39"] = [0 for i in range(5)]
            ws["J35"], ws["J36"], ws["J37"], ws["J38"], ws["J39"] = [0 for i in range(5)]
            ws["K35"], ws["K36"], ws["K37"], ws["K38"], ws["K39"] = [0 for i in range(5)]
            ws["L35"], ws["L36"], ws["L37"], ws["L38"], ws["L39"] = [0 for i in range(5)]
            ws["M35"], ws["M36"], ws["M37"], ws["M38"], ws["M39"] = [0 for i in range(5)]
            ws["N35"], ws["N36"], ws["N37"], ws["N38"], ws["N39"] = [0 for i in range(5)]
            ws["O35"], ws["O36"], ws["O37"], ws["O38"], ws["O39"] = [0 for i in range(5)]

        time.sleep(3)

        # Get Monthly Check Out Statistics
        click(stat_btn)
        send_checkout_date(monthly_stat_from, loan_date_to)

        try:
            search_and_close()
            get_page_info()  # Parsing

            # Save in Excel
            ws["F41"] = get_page_info()["monthly_general_works"]
            ws["G41"] = get_page_info()["monthly_philosophy"]
            ws["H41"] = get_page_info()["monthly_religion"]
            ws["I41"] = get_page_info()["monthly_social_sciences"]
            ws["J41"] = get_page_info()["monthly_natural_sciences"]
            ws["K41"] = get_page_info()["monthly_technology"]
            ws["L41"] = get_page_info()["monthly_arts"]
            ws["M41"] = get_page_info()["monthly_language"]
            ws["N41"] = get_page_info()["monthly_literature"]
            ws["O41"] = get_page_info()["monthly_history"]

        except IndexError:
            # Save in Excel
            ws["F41"] = 0
            ws["G41"] = 0
            ws["H41"] = 0
            ws["I41"] = 0
            ws["J41"] = 0
            ws["K41"] = 0
            ws["L41"] = 0
            ws["M41"] = 0
            ws["N41"] = 0
            ws["O41"] = 0

        time.sleep(3)

        # Get Return Statistics
        driver.find_element_by_xpath('//*[@id="statistics_type"]/option[8]').click()
        click(stat_btn)
        time.sleep(1)
        uncheck_checkout = driver.find_element_by_xpath(
            '//*[@id="loan_return_statistics_search_modal"]/div/div/div[2]/div[1]/table[2]/tbody/tr[1]/th[2]/label')
        click(uncheck_checkout)
        send_return_date(loan_date_from, loan_date_to)
        driver.implicitly_wait(10)
        by_date_2 = driver.find_element_by_xpath('//*[@id="jqg162"]/td[2]')
        click(by_date_2)

        try:
            search_and_close()
            get_page_info()  # Parsing

            # Save in Excel
            ws["F46"], ws["F47"], ws["F48"], ws["F49"], ws["F50"] = [get_page_info()["general_works"][i] for i in
                                                                     range(5)]
            ws["G46"], ws["G47"], ws["G48"], ws["G49"], ws["G50"] = [get_page_info()["philosophy"][i] for i in range(5)]
            ws["H46"], ws["H47"], ws["H48"], ws["H49"], ws["H50"] = [get_page_info()["religion"][i] for i in range(5)]
            ws["I46"], ws["I47"], ws["I48"], ws["I49"], ws["I50"] = [get_page_info()["social_sciences"][i] for i in
                                                                     range(5)]
            ws["J46"], ws["J47"], ws["J48"], ws["J49"], ws["J50"] = [get_page_info()["natural_sciences"][i] for i in
                                                                     range(5)]
            ws["K46"], ws["K47"], ws["K48"], ws["K49"], ws["K50"] = [get_page_info()["technology"][i] for i in range(5)]
            ws["L46"], ws["L47"], ws["L48"], ws["L49"], ws["L50"] = [get_page_info()["arts"][i] for i in range(5)]
            ws["M46"], ws["M47"], ws["M48"], ws["M49"], ws["M50"] = [get_page_info()["language"][i] for i in range(5)]
            ws["N46"], ws["N47"], ws["N48"], ws["N49"], ws["N50"] = [get_page_info()["literature"][i] for i in range(5)]
            ws["O46"], ws["O47"], ws["O48"], ws["O49"], ws["O50"] = [get_page_info()["history"][i] for i in range(5)]

        except IndexError:
            # Save in Excel
            ws["F46"], ws["F47"], ws["F48"], ws["F49"], ws["F50"] = [0 for i in range(5)]
            ws["G46"], ws["G47"], ws["G48"], ws["G49"], ws["G50"] = [0 for i in range(5)]
            ws["H46"], ws["H47"], ws["H48"], ws["H49"], ws["H50"] = [0 for i in range(5)]
            ws["I46"], ws["I47"], ws["I48"], ws["I49"], ws["I50"] = [0 for i in range(5)]
            ws["J46"], ws["J47"], ws["J48"], ws["J49"], ws["J50"] = [0 for i in range(5)]
            ws["K46"], ws["K47"], ws["K48"], ws["K49"], ws["K50"] = [0 for i in range(5)]
            ws["L46"], ws["L47"], ws["L48"], ws["L49"], ws["L50"] = [0 for i in range(5)]
            ws["M46"], ws["M47"], ws["M48"], ws["M49"], ws["M50"] = [0 for i in range(5)]
            ws["N46"], ws["N47"], ws["N48"], ws["N49"], ws["N50"] = [0 for i in range(5)]
            ws["O46"], ws["O47"], ws["O48"], ws["O49"], ws["O50"] = [0 for i in range(5)]

        time.sleep(3)

        # Get Monthly Return Statistics
        click(stat_btn)
        time.sleep(1)
        send_return_date(monthly_stat_from, loan_date_to)

        try:
            search_and_close()
            time.sleep(3)
            get_page_info()  # Parsing

            # Save in Excel
            ws["F52"] = get_page_info()["monthly_general_works"]
            ws["G52"] = get_page_info()["monthly_philosophy"]
            ws["H52"] = get_page_info()["monthly_religion"]
            ws["I52"] = get_page_info()["monthly_social_sciences"]
            ws["J52"] = get_page_info()["monthly_natural_sciences"]
            ws["K52"] = get_page_info()["monthly_technology"]
            ws["L52"] = get_page_info()["monthly_arts"]
            ws["M52"] = get_page_info()["monthly_language"]
            ws["N52"] = get_page_info()["monthly_literature"]
            ws["O52"] = get_page_info()["monthly_history"]

        except IndexError:
            # Save in Excel
            ws["F52"] = 0
            ws["G52"] = 0
            ws["H52"] = 0
            ws["I52"] = 0
            ws["J52"] = 0
            ws["K52"] = 0
            ws["L52"] = 0
            ws["M52"] = 0
            ws["N52"] = 0
            ws["O52"] = 0

        time.sleep(3)

        # Get Return User Statistics
        driver.find_element_by_xpath('//*[@id="statistics_type"]/option[2]').click()
        click(stat_btn)
        time.sleep(1)
        register_date = driver.find_element_by_xpath('//*[@id="reg_date_from_text"]')
        click(register_date)
        expected_return_date = driver.find_element_by_xpath('//*[@id="loan_return_statistics_search_modal"]/div/div/div[2]/div[1]/table[2]/tbody/tr[2]/th[1]/label')
        click(expected_return_date)
        send_return_date(loan_date_from, loan_date_to)
        driver.implicitly_wait(10)
        by_date_3 = driver.find_element_by_xpath('//*[@id="jqg240"]/td[2]')
        click(by_date_3)
        by_class = driver.find_element_by_xpath('//*[@id="jqg264"]/td[2]')
        click(by_class)

        try:
            search_and_close()
            time.sleep(3)

            # Save in Excel
            ws["D46"], ws["D47"], ws["D48"], ws["D49"], ws["D50"] = [get_user_statistics()["korean"][i] for i in range(5)]
            ws["E46"], ws["E47"], ws["E48"], ws["E49"], ws["E50"] = [get_user_statistics()["foreigner"][i] for i in range(5)]

        except:
            # Save in Excel
            ws["D46"], ws["D47"], ws["D48"], ws["D49"], ws["D50"] = [0 for i in range(5)]
            ws["E46"], ws["E47"], ws["E48"], ws["E49"], ws["E50"] = [0 for i in range(5)]

        time.sleep(3)

        # Get Monthly Return User Statistics
        click(stat_btn)
        time.sleep(1)
        send_return_date(monthly_stat_from, loan_date_to)

        try:
            search_and_close()
            time.sleep(3)

            # Save in Excel
            ws["D52"] = get_user_statistics()["monthly_korean"]
            ws["E52"] = get_user_statistics()["monthly_foreigner"]

        except IndexError:
            # Save in Excel
            ws["D52"] = 0
            ws["E52"] = 0

        # Get Check Out User Statistics
        driver.find_element_by_xpath('//*[@id="statistics_type"]/option[1]').click()
        click(stat_btn)
        time.sleep(1)
        send_checkout_date(loan_date_from, loan_date_to)
        by_date_4 = driver.find_element_by_xpath('//*[@id="jqg342"]/td[2]')
        click(by_date_4)
        by_class_2 = driver.find_element_by_xpath('//*[@id="jqg364"]/td[2]')
        click(by_class_2)

        try:
            search_and_close()
            time.sleep(3)

            # Save in Excel
            ws["F11"], ws["F12"], ws["F13"], ws["F14"], ws["F15"] = [get_user_statistics()["korean"][i] for i in
                                                                     range(5)]
            ws["G11"], ws["G12"], ws["G13"], ws["G14"], ws["G15"] = [get_user_statistics()["foreigner"][i] for i in
                                                                     range(5)]

        except IndexError:
            # Save in Excel
            ws["F11"], ws["F12"], ws["F13"], ws["F14"], ws["F15"] = [0 for i in range(5)]
            ws["G11"], ws["G12"], ws["G13"], ws["G14"], ws["G15"] = [0 for i in range(5)]
            ws["H11"], ws["I11"], ws["H12"], ws["I12"], ws["H13"], ws["I13"], ws["H14"], ws["I14"], ws["H15"], ws[
                "I15"] = [0 for i in range(10)]

        time.sleep(3)

        # Get Monthly Check Out User Statistics
        click(stat_btn)
        time.sleep(1)
        send_checkout_date(monthly_stat_from, loan_date_to)

        try:
            search_and_close()
            time.sleep(3)

            # Save in Excel
            ws["F17"] = get_user_statistics()["monthly_korean"]
            ws["G17"] = get_user_statistics()["monthly_foreigner"]

        except IndexError:
            # Save in Excel
            ws["F17"] = 0
            ws["G17"] = 0

        time.sleep(3)

        # Get New User Statistics
        driver.find_element_by_xpath('//*[@id="statistics_type"]/option[3]').click()
        click(stat_btn)
        time.sleep(1)
        send_register_date(loan_date_from, loan_date_to)
        reset = driver.find_element_by_xpath('//*[@id="shelf_loc_code_close_btn"]')
        click(reset)
        driver.implicitly_wait(10)
        by_date_5 = driver.find_element_by_xpath('//*[@id="jqg442"]/td[2]')
        click(by_date_5)
        by_class_3 = driver.find_element_by_xpath('//*[@id="jqg461"]/td[2]')
        click(by_class_3)

        try:
            search_and_close()
            # Save in Excel
            ws["G23"], ws["G24"], ws["G25"], ws["G26"], ws["G27"] = [get_user_statistics()["korean"][i] for i in
                                                                     range(5)]
            ws["K23"], ws["K24"], ws["K25"], ws["K26"], ws["K27"] = [get_user_statistics()["foreigner"][i] for i in
                                                                     range(5)]

        except IndexError:
            ws["G23"], ws["G24"], ws["G25"], ws["G26"], ws["G27"] = [0 for i in range(5)]
            ws["K23"], ws["K24"], ws["K25"], ws["K26"], ws["K27"] = [0 for i in range(5)]

        time.sleep(3)

        # Get Monthly New User Statistics
        click(stat_btn)
        time.sleep(1)
        send_register_date(monthly_stat_from, loan_date_to)

        try:
            search_and_close()
            # Save in Excel
            ws["G29"] = get_user_statistics()["monthly_korean"]
            ws["K29"] = get_user_statistics()["monthly_foreigner"]

        except IndexError:
            ws["G29"] = 0
            ws["K29"] = 0

        time.sleep(3)

        # Get Collection Statistics
        menu_2 = driver.find_element_by_xpath('//*[@id="right_container_wrapper"]/header/div[5]/nav/a')  # Open Statistics tap
        click(menu_2)
        collection_stat = driver.find_element_by_xpath('//*[@id="MOLI_02_HTML"]/li[14]')
        click(collection_stat)
        driver.switch_to.window(driver.window_handles[2])
        driver.find_element_by_xpath('//*[@id="statistics_type"]/option[6]').click()
        book_count = driver.find_element_by_xpath('//*[@id="non_price_type"]/label[2]')
        click(book_count)
        remove_zero = driver.find_element_by_xpath('//*[@id="right_container_wrapper"]/main/div/div[1]/form/div[2]/label[4]')
        click(remove_zero)
        stat_btn_2 = driver.find_element_by_xpath('//*[@id="statistics_btn"]')
        click(stat_btn_2)
        time.sleep(1)
        send_shelf_date('2013/01/01', loan_date_to)
        checkbox2 = driver.find_element_by_xpath('//*[@id="shelf_hold_statistics_search_modal"]/div/div/div[2]/div[1]/table/tbody/tr[4]/th[1]/label')
        click(checkbox2)
        checkbox3 = driver.find_element_by_xpath('//*[@id="shelf_hold_statistics_search_modal"]/div/div/div[2]/div[1]/table/tbody/tr[5]/th[1]/label')
        click(checkbox3)
        driver.find_element_by_xpath('//*[@id="working_status"]/option[4]').click()
        search_and_close()
        driver.implicitly_wait(120)

        # Save in Excel
        ws4["B4"], ws4["C4"], ws4["D4"], ws4["E4"], ws4["F4"], ws4["G4"], ws4["H4"], ws4["I4"], ws4["J4"], ws4["K4"], \
        ws4["L4"] = [get_collection_statistics()["korean_big"][i] for i in range(11)]
        ws4["B5"], ws4["C5"], ws4["D5"], ws4["E5"], ws4["F5"], ws4["G5"], ws4["H5"], ws4["I5"], ws4["J5"], ws4["K5"], \
        ws4["L5"] = [get_collection_statistics()["korean"][i] for i in range(11)]
        ws4["B6"], ws4["C6"], ws4["D6"], ws4["E6"], ws4["F6"], ws4["G6"], ws4["H6"], ws4["I6"], ws4["J6"], ws4["K6"], \
        ws4["L6"] = [get_collection_statistics()["korean_refer"][i] for i in range(11)]
        ws4["B7"], ws4["C7"], ws4["D7"], ws4["E7"], ws4["F7"], ws4["G7"], ws4["H7"], ws4["I7"], ws4["J7"], ws4["K7"], \
        ws4["L7"] = [get_collection_statistics()["non_book"][i] for i in range(11)]
        ws4["B8"], ws4["C8"], ws4["D8"], ws4["E8"], ws4["F8"], ws4["G8"], ws4["H8"], ws4["I8"], ws4["J8"], ws4["K8"], \
        ws4["L8"] = [get_collection_statistics()["chinese"][i] for i in range(11)]
        ws4["B9"], ws4["C9"], ws4["D9"], ws4["E9"], ws4["F9"], ws4["G9"], ws4["H9"], ws4["I9"], ws4["J9"], ws4["K9"], \
        ws4["L9"] = [get_collection_statistics()["french"][i] for i in range(11)]
        ws4["B10"], ws4["C10"], ws4["D10"], ws4["E10"], ws4["F10"], ws4["G10"], ws4["H10"], ws4["I10"], ws4["J10"], ws4["K10"], ws4["L10"]\
            = [get_collection_statistics()["english"][i] for i in range(11)]
        ws4["B11"], ws4["C11"], ws4["D11"], ws4["E11"], ws4["F11"], ws4["G11"], ws4["H11"], ws4["I11"], ws4["J11"], ws4["K11"], ws4["L11"]\
            = [get_collection_statistics()["japanese"][i] for i in range(11)]
        ws4["B12"], ws4["C12"], ws4["D12"], ws4["E12"], ws4["F12"], ws4["G12"], ws4["H12"], ws4["I12"], ws4["J12"], ws4["K12"], ws4["L12"]\
            = [get_collection_statistics()["mongolian"][i] for i in range(11)]
        ws4["B13"], ws4["C13"], ws4["D13"], ws4["E13"], ws4["F13"], ws4["G13"], ws4["H13"], ws4["I13"], ws4["J13"], ws4["K13"], ws4["L13"]\
            = [get_collection_statistics()["filipino"][i] for i in range(11)]
        ws4["B14"], ws4["C14"], ws4["D14"], ws4["E14"], ws4["F14"], ws4["G14"], ws4["H14"], ws4["I14"], ws4["J14"], ws4["K14"], ws4["L14"]\
            = [get_collection_statistics()["english_refer"][i] for i in range(11)]
        ws4["B15"], ws4["C15"], ws4["D15"], ws4["E15"], ws4["F15"], ws4["G15"], ws4["H15"], ws4["I15"], ws4["J15"], ws4["K15"], ws4["L15"]\
            = [get_collection_statistics()["spanish"][i] for i in range(11)]
        ws4["B16"], ws4["C16"], ws4["D16"], ws4["E16"], ws4["F16"], ws4["G16"], ws4["H16"], ws4["I16"], ws4["J16"], ws4["K16"], ws4["L16"]\
            = [get_collection_statistics()["vietnamese"][i] for i in range(11)]
        ws4["B17"], ws4["C17"], ws4["D17"], ws4["E17"], ws4["F17"], ws4["G17"], ws4["H17"], ws4["I17"], ws4["J17"], ws4["K17"], ws4["L17"]\
            = [get_collection_statistics()["cambodian"][i] for i in range(11)]

        # Get Purchase Statistics
        purchase_row_list = [4, 5, 6, 7, 8, 9, 10, 11, 12, 13, 14, 15]
        donation_row_list = [21, 22, 23, 24, 25, 26, 27, 28, 29, 30, 31, 32]
        purchase_row = purchase_row_list[int(loan_date_from_dt.month) - 1]
        donation_row = donation_row_list[int(loan_date_from_dt.month) - 1]

        click(stat_btn_2)
        time.sleep(1)
        send_shelf_date(monthly_stat_from, loan_date_to)
        driver.find_element_by_xpath('//*[@id="acq_code"]/option[2]').click()

        try:
            search_and_close()
            driver.implicitly_wait(120)
            ws3[f"B{purchase_row}"], ws3[f"C{purchase_row}"], ws3[f"D{purchase_row}"], ws3[f"E{purchase_row}"], ws3[
                f"F{purchase_row}"], ws3[f"G{purchase_row}"], ws3[f"H{purchase_row}"], \
            ws3[f"I{purchase_row}"], ws3[f"J{purchase_row}"], ws3[f"K{purchase_row}"], ws3[f"L{purchase_row}"] = [get_purchase_statistics()["sum"][i] for i in range(11)]

        except IndexError:
            ws3[f"B{purchase_row}"], ws3[f"C{purchase_row}"], ws3[f"D{purchase_row}"], ws3[f"E{purchase_row}"], ws3[
                f"F{purchase_row}"], ws3[f"G{purchase_row}"], ws3[f"H{purchase_row}"], \
            ws3[f"I{purchase_row}"], ws3[f"J{purchase_row}"], ws3[f"K{purchase_row}"], ws3[f"L{purchase_row}"] = [0 for i in range(11)]

        # Get Donation Statistics
        click(stat_btn_2)
        time.sleep(1)
        driver.find_element_by_xpath('//*[@id="acq_code"]/option[3]').click()

        try:
            search_and_close()
            driver.implicitly_wait(120)
            ws3[f"B{donation_row}"], ws3[f"C{donation_row}"], ws3[f"D{donation_row}"], ws3[f"E{donation_row}"], ws3[
                f"F{donation_row}"], ws3[f"G{donation_row}"], ws3[f"H{donation_row}"], \
            ws3[f"I{donation_row}"], ws3[f"J{donation_row}"], ws3[f"K{donation_row}"], ws3[f"L{donation_row}"] = [get_donation_statistics()["sum"][i] for i in range(11)]

        except IndexError:
            ws3[f"B{donation_row}"], ws3[f"C{donation_row}"], ws3[f"D{donation_row}"], ws3[f"E{donation_row}"], ws3[
                f"F{donation_row}"], ws3[f"G{donation_row}"], ws3[f"H{donation_row}"], \
            ws3[f"I{donation_row}"], ws3[f"J{donation_row}"], ws3[f"K{donation_row}"], ws3[f"L{donation_row}"] = [0 for i in range(11)]

        # Get Dates
        monday = loan_date_from_dt
        tuesday = loan_date_from_dt + timedelta(days=1)
        wednesday = loan_date_from_dt + timedelta(days=2)
        thursday = loan_date_from_dt + timedelta(days=3)
        friday = loan_date_from_dt + timedelta(days=4)
        saturday = loan_date_from_dt + timedelta(days=5)
        sunday = loan_date_from_dt + timedelta(days=6)

        # Save in Excel
        ws["A1"] = "일별 이용실적 통계\n" + loan_date_from + "(월) ~ " + loan_date_to + "(금)"
        ws4["A2"] = f"단위: 권                                                                                                                                                                                                            {loan_date_to} 기준"
        ws3["A2"] = f"단위:권                                                                                                                      {loan_date_to} 기준"
        ws3["A19"] = f"단위:권                                                                                                                      {loan_date_to} 기준"

        ws["A11"], ws["A23"], ws["A35"], ws["A46"] = [str(monday.month) + "/" + str(monday.day) for i in range(0, 4)]
        ws["A12"], ws["A24"], ws["A36"], ws["A47"] = [str(tuesday.month) + "/" + str(tuesday.day) for i in range(0, 4)]
        ws["A13"], ws["A25"], ws["A37"], ws["A48"] = [str(wednesday.month) + "/" + str(wednesday.day) for i in range(0, 4)]
        ws["A14"], ws["A26"], ws["A38"], ws["A49"] = [str(thursday.month) + "/" + str(thursday.day) for i in range(0, 4)]
        ws["A15"], ws["A27"], ws["A39"], ws["A50"] = [str(friday.month) + "/" + str(friday.day) for i in range(0, 4)]

        ws["B4"] = str(monday.month) + "/" + str(monday.day) + "(월)"
        ws["E4"] = str(tuesday.month) + "/" + str(tuesday.day) + "(화)"
        ws["H4"] = str(wednesday.month) + "/" + str(wednesday.day) + "(수)"
        ws["K4"] = str(thursday.month) + "/" + str(thursday.day) + "(목)"
        ws["N4"] = str(friday.month) + "/" + str(friday.day) + "(금)"

        wb.save(os.path.dirname(os.path.abspath(__file__)) + f"\{self.file_name_original.text()}_new.xlsx")

if __name__ == "__main__":
    import sys
    app = QtWidgets.QApplication(sys.argv)
    Alpas = QtWidgets.QDialog()
    ui = Ui_Alpas()
    ui.setupUi(Alpas)
    Alpas.show()
    sys.exit(app.exec_())
